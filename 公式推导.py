import math
import random
import sympy as sp
from sympy import *
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

fm, t0, tm, r, d, t = sp.symbols('fm t0 tm r d t')
f1 = fm
f2 = ((t + t0)/(tm + t0))**r
f31 = d/(d + r)
f321 = r/(d + r)
f322 = ((t + t0)/(tm + t0))**(r + 1)
f32 = (f321 * f322)
n = -(r + d)/(r + 1)
f3 = (f31 + f32)**n
fuc = f1 * f2 * f3
fuc_x = sp.diff(fuc, t)
# print("对x求导结果 : ", fuc)

def hu(fm, t0, tm, r, d):
    x = fm*r*((t + t0)/(t0 + tm))**r*((t + t0)/(t0 + tm))**(r + 1)*(-d - r)*(d/(d + r) + r*((t + t0)/(t0 + tm))**(r + 1)/(d + r))**((-d - r)/(r + 1))/((d + r)*(t + t0)*(d/(d + r) + r*((t + t0)/(t0 + tm))**(r + 1)/(d + r))) + fm*r*((t + t0)/(t0 + tm))**r*(d/(d + r) + r*((t + t0)/(t0 + tm))**(r + 1)/(d + r))**((-d - r)/(r + 1))/(t + t0)
    re = solve(x, t)
    n = len(re)
    v = [0 for _ in range(n)]
    for i in range(0, n):
        v[i] = fm * ((re[i] + t0) / (t0 + tm)) ** r * (d / (d + r) + r * ((re[i] + t0) / (t0 + tm)) ** (r + 1) / (d + r)) ** ((-d - r) / (r + 1))
        if v[i] == max(v):
            max_y = v[i]
            max_x = re[i]
    x1 = fm * ((t + t0) / (t0 + tm)) ** r * (d / (d + r) + r * ((t + t0) / (t0 + tm)) ** (r + 1) / (d + r)) ** ((-d - r) / (r + 1)) - max_y * 0.5
    re1 = solve(x1, t)
    bfz = max_y * 0.5

    fh = [0 for _ in range(5)]
    fh[0] = max_x
    fh[1] = max_y
    fh[2] = re1[0]
    fh[3] = re1[1]
    fh[4] = bfz
    return fh


wb = openpyxl.load_workbook('1.xlsx') # 打开excel文件
sheet = wb.active  # 正对表格

n1 = sheet.max_row
n2 = sheet.max_column

data = [[] for _ in range(n1)]
list = [[] for _ in range(n2)]

for column in range(1, n1+1):
    for row in range(1, n2+1):
        list[row-1] = sheet.cell(column, row).value
    data[column-1]=list
    list = [[] for _ in range(n2)]


# 'tm', 'fm', 't0', 'r', 'd'
# def hu(fm, t0, tm, r, d):

data1 = [[] for _ in range(n1)]
data1[0] = ['峰值对应的x','峰值','左解x1','右解x2','半峰值']
for i in range(0, n1-1):
    print(data1[i])
    data1[i+1] = hu(data[i+1][4], data[i+1][5], data[i+1][3], data[i+1][6], data[i+1][7])
    # print(data1[i])
wb = Workbook()
ws = wb.active
for i in range(0, n1):
    for j in range(0, 5):
        ws.cell(row=j + 1, column=i + 1).value = data1[i][j]


wb.save("短暴拟合.xlsx") #保存

# wb1 = openpyxl.load_workbook('1.xlsx') # 打开excel文件
# sheet1 = wb1.active  # 正对表格
#
# n11 = sheet1.max_row
# n22 = sheet1.max_column
#
# data1 = [[] for _ in range(n11)]
# list1 = [[] for _ in range(n22)]
#
# for column in range(1,n11+1):
#     for row in range(1,n22+1):
#         list1[row-1] = sheet1.cell(column, row).value
#     data1[column-1]=list1
#     list1 = [[] for _ in range(n22)]
#
# # =========合并单元格区===========
#
# m1 = 0
# m2 = 0
#
# for j in range(1, n11):
#     v = str(data1[j][0])
#     if v == '短暴':
#         m1 = m1 + 1
#     if v == '长暴':
#         m2 = m2 + 1
#
# # 定义合并单元格的开始位置
# for j1 in range(1, m1+1):
#     ws.merge_cells(start_row=j1+1, end_row=m1+1, start_column=1, end_column=1)
# # for j2 in range(m1+1, n11+1):
#     ws.merge_cells(start_row=m1+1, end_row=n11, start_column=1, end_column=1)
#
#
# print(m1,m2,n11)
# # 工作簿保存到磁盘
# wb.save('test1.xlsx')









